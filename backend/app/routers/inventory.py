from typing import List
from io import BytesIO
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, status
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app import models, schemas
from app.routers.auth import require_roles, is_privileged


router = APIRouter()


@router.get("/", response_model=List[schemas.InventoryRead])
def list_inventory(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(
        require_roles("Admin", "BuyerRep")
    ),
):
    query = db.query(models.Inventory)

    # Location: everyone except Admin & Finance is limited by location
    if not is_privileged(current_user):
        query = query.filter(models.Inventory.location == current_user.location)

    # PR: cannot see sold / under-contract cars in inventory
    if current_user.role == "PR":
        query = query.filter(
            ~models.Inventory.status.in_(["Under Contract", "Sold"])
        )

    return query.all()


@router.post("/", response_model=schemas.InventoryRead)
def create_inventory(
    item: schemas.InventoryCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_roles("Admin", "BuyerRep")),
):
    """
    Create a single inventory record.

    Rules:
    - Non-admins can only create in their own location.
    - Car age ≤ 25 years.
    - Mileage < 150000.
    - Profit% = (sale_price - cost) / cost * 100.
      * Admin: profit ≥ 5%
      * BuyerRep: profit ≥ 35%
    - Status is set automatically:
      * Non-damaged → 'Available'
      * Damaged → 'In Service' and also auto-added to Service table.
    """

    # Non-admins can only create in their own location
    if current_user.role == "BuyerRep":
        item.location = current_user.location

    # Validate age
    current_year = datetime.now().year
    age = current_year - item.year
    if age > 25:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Car age exceeds 25 years limit.",
        )

    # Validate mileage
    if item.mileage >= 150000:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Mileage must be less than 150,000.",
        )

    # Validate cost
    if item.cost <= 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cost must be greater than 0.",
        )

    # Compute profit percentage
    profit_percent = ((item.sale_price - item.cost) / item.cost) * 100.0

    # Enforce min profit: Admin ≥ 5%, BuyerRep ≥ 35%
    min_profit = 5.0 if current_user.role == "Admin" else ( 21.5 if current_user.role == "BuyerRep" else 35.0)
    if profit_percent < min_profit:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Profit below minimum threshold ({min_profit}%).",
        )

    # Determine status
    if item.condition_type == "Damaged":
        status_value = "In Service"  # Not available, immediately in service
    else:
        status_value = "Available"

    # Auto-generate VIN if not provided
    if not item.vin_number:
        now = datetime.now()
        # Format: MMYYYYN – for simplicity, N=1 here. You can enhance later with a real sequence.
        item.vin_number = f"{now.month:02d}{now.year}{1}"

    # Create inventory record
    # Create inventory record WITHOUT VIN first
    db_item = models.Inventory(
        vin_number="",
        make=item.make,
        model=item.model,
        year=item.year,
        mileage=item.mileage,
        condition_type=item.condition_type,
        cost=item.cost,
        sale_price=item.sale_price,
        profit_percent=profit_percent,
        status=status_value,
        location=item.location,
    )
    db.add(db_item)
    db.commit()
    db.refresh(db_item)

    # Now generate VIN based on date + row number (id)
    now = datetime.now()
    generated_vin = f"{now.month:02d}{now.day:02d}{now.year}{db_item.id}"
    db_item.vin_number = generated_vin
    db.commit()
    db.refresh(db_item)

    # Use generated_vin for Service if Damaged
    if item.condition_type == "Damaged":
        service_id = f"{now.month:02d}{now.day:02d}{now.year}{db_item.id}"
        service_record = models.Service(
            service_id=service_id,
            vin_number=generated_vin,
            seriousness_level="High",
            estimated_days=3,
            cost_added=2000,
            status="In Service",
        )
        db.add(service_record)
        db.commit()

    return db_item


# ... keep your other imports at the top
@router.post("/upload")
async def upload_inventory_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_roles("Admin", "BuyerRep")),
):
    """
    Excel upload endpoint for batch inventory import.

    Expected header row (case-insensitive):

    Make | Model | Year | Mileage | Condition | Cost | Sale Price | Location

    - VIN is NOT in the file; we generate it as MMDDYYYY<id>.
    - profit_percent and status are computed automatically.
    - Damaged cars are auto-marked 'In Service' and added to service table.
    - Admin: location is taken from Excel 'Location' column (required).
    - BuyerRep: only rows whose Location matches current_user.location are imported;
      others are rejected with an error message.
    """
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No file uploaded.",
        )

    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only .xlsx/.xls files are supported.",
        )

    # Read whole file into memory
    raw_bytes = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(raw_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Could not read Excel file: {e}",
        )

    ws = wb.active  # first sheet

    # Read header row
    headers = [cell.value for cell in ws[1]]
    if not headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Excel file has no header row.",
        )

    # Normalize header names (strip spaces, lower)
    normalized = [str(h).strip().lower() if h is not None else "" for h in headers]
    col_map = {name: idx for idx, name in enumerate(normalized)}

    # Map your actual columns
    required_cols = [
        "make",
        "model",
        "year",
        "mileage",
        "condition",    # maps to condition_type
        "cost",
        "sale price",   # maps to sale_price
        "location",     # NEW: must come from Excel
    ]

    for col in required_cols:
        if col not in col_map:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Missing required column in Excel header: {col}",
            )

    imported = 0
    errors: list[str] = []
    current_year = datetime.now().year

    for row_idx in range(2, ws.max_row + 1):
        try:
            def get(col_name: str):
                col_index = col_map[col_name] + 1  # openpyxl is 1-based
                return ws.cell(row=row_idx, column=col_index).value

            make = str(get("make") or "").strip()
            model = str(get("model") or "").strip()
            year_val = get("year")
            mileage_val = get("mileage")
            condition_type = str(get("condition") or "").strip()
            cost_val = get("cost")
            sale_price_val = get("sale price")
            excel_location = str(get("location") or "").strip()

            if not make or not model:
                raise ValueError("Make and Model are required.")

            if year_val is None or mileage_val is None:
                raise ValueError("Year and Mileage are required.")

            if condition_type == "":
                raise ValueError("Condition is required.")

            if cost_val is None or sale_price_val is None:
                raise ValueError("Cost and Sale Price are required.")

            # Normalize numeric fields
            year = int(year_val)
            mileage = int(mileage_val)
            cost = float(cost_val)
            sale_price = float(sale_price_val)

            # LOCATION RULES
            if current_user.role == "Admin":
                # Admin: location must come from Excel
                if not excel_location:
                    raise ValueError(
                        "Location is required in Excel for admin import."
                    )
                location = excel_location
            else:  # BuyerRep
                if not excel_location:
                    raise ValueError(
                        "Location is required in Excel for buyer rep imports."
                    )
                if excel_location != current_user.location:
                    # Reject row: wrong location
                    raise ValueError(
                        f"location '{excel_location}' does not match buyer rep location '{current_user.location}'"
                    )
                location = excel_location

            # Age rule
            age = current_year - year
            if age > 25:
                raise ValueError("Car age exceeds 25 years limit.")

            # Mileage rule
            if mileage >= 150000:
                raise ValueError("Mileage must be less than 150,000.")

            if cost <= 0:
                raise ValueError("Cost must be greater than 0.")

            # Profit rules
            profit_percent = ((sale_price - cost) / cost) * 100.0
            min_profit = 5.0 if current_user.role == "Admin" else 35.0
            if profit_percent < min_profit:
                raise ValueError(f"Profit below minimum threshold ({min_profit}%).")

            # Determine status
            if condition_type.lower() == "damaged":
                status_value = "In Service"
            else:
                status_value = "Available"

            # First create inventory without VIN to get id
            inv = models.Inventory(
                vin_number="",
                make=make,
                model=model,
                year=year,
                mileage=mileage,
                condition_type=condition_type,
                cost=cost,
                sale_price=sale_price,
                profit_percent=profit_percent,
                status=status_value,
                location=location,
            )
            db.add(inv)
            db.flush()  # get inv.id

            # Now generate VIN: MMDDYYYY<id>
            now = datetime.now()
            generated_vin = f"{now.month:02d}{now.day:02d}{now.year}{inv.id}"
            inv.vin_number = generated_vin

            # If Damaged, create Service record
            if condition_type.lower() == "damaged":
                service_id = f"{now.month:02d}{now.day:02d}{now.year}{inv.id}"
                service_record = models.Service(
                    service_id=service_id,
                    vin_number=generated_vin,
                    seriousness_level="High",
                    estimated_days=3,
                    cost_added=2000,
                    status="In Service",
                )
                db.add(service_record)

            imported += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {e}")

    db.commit()

    detail_msg = f"Imported {imported} cars from Excel."
    if errors:
        detail_msg += f" Some rows failed: {len(errors)}. First error: {errors[0]}"

    return {"detail": detail_msg}



@router.get("/{vin}", response_model=schemas.InventoryRead)
def get_inventory(
    vin: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(
        require_roles("Admin", "SalesRep", "BuyerRep")
    ),
):
    item = db.query(models.Inventory).filter(models.Inventory.vin_number == vin).first()
    if not item:
        raise HTTPException(status_code=404, detail="Car not found")

    if not is_privileged(current_user) and item.location != current_user.location:
        raise HTTPException(status_code=403, detail="Not enough permissions")

    return item


@router.delete("/{vin}")
def delete_inventory(
    vin: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_roles("Admin")),
):
    item = db.query(models.Inventory).filter(models.Inventory.vin_number == vin).first()
    if not item:
        raise HTTPException(status_code=404, detail="Car not found")

    db.delete(item)
    db.commit()
    return {"detail": "Deleted"}
@router.patch("/{vin}", response_model=schemas.InventoryRead)
def update_inventory(
    vin: str,
    update: schemas.InventoryUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(require_roles("Admin", "BuyerRep")),
):
    """
    Update one or more properties of a single car in inventory.

    - Admin: can edit any car in any location.
    - BuyerRep: only cars in their own location.
    - Enforces age, mileage, and profit rules.
    - Damaged condition auto-moves car to Service with status=In Service.
    """
    item = (
        db.query(models.Inventory)
        .filter(models.Inventory.vin_number == vin)
        .first()
    )
    if not item:
        raise HTTPException(status_code=404, detail="Car not found")

    # BuyerRep location restriction
    if not is_privileged(current_user) and item.location != current_user.location:
        raise HTTPException(
            status_code=403,
            detail="Not enough permissions for this location",
        )

    data = update.dict(exclude_unset=True)
    # Apply basic field updates in-memory
    for field, value in data.items():
        if hasattr(item, field) and field not in ("id", "vin_number"):
            setattr(item, field, value)

    # Now enforce business rules using updated values
    now = datetime.now()
    current_year = now.year

    new_year = item.year
    new_mileage = item.mileage
    new_cost = item.cost
    new_sale_price = item.sale_price
    new_condition = item.condition_type

    # Age rule
    age = current_year - new_year
    if age > 25:
        raise HTTPException(
            status_code=400,
            detail="Car age exceeds 25 years limit.",
        )

    # Mileage rule
    if new_mileage >= 150000:
        raise HTTPException(
            status_code=400,
            detail="Mileage must be less than 150,000.",
        )

    if new_cost <= 0:
        raise HTTPException(
            status_code=400,
            detail="Cost must be greater than 0.",
        )

    # Profit rule
    profit_percent = ((new_sale_price - new_cost) / new_cost) * 100.0

    if current_user.role == "Admin":
        min_profit = 5.0
    elif current_user.role == "BuyerRep":
        # BuyerRep discount limit: equivalent of 10% off a 35% base profit
        min_profit = 21.5
    else:
        # default safeguard if role ever expanded
        min_profit = 35.0

    if profit_percent < min_profit:
        raise HTTPException(
            status_code=400,
            detail=f"Profit below minimum threshold ({min_profit}%).",
        )

    item.profit_percent = profit_percent

    # Condition / Service rule
    if new_condition.lower() == "damaged":
        # Mark inventory as in service
        item.status = "In Service"

        # Ensure a Service record exists
        existing_service = (
            db.query(models.Service)
            .filter(models.Service.vin_number == item.vin_number,
                    models.Service.status == "In Service")
            .first()
        )
        if not existing_service:
            service_id = f"{now.month:02d}{now.day:02d}{now.year}{item.id}"
            service_record = models.Service(
                service_id=service_id,
                vin_number=item.vin_number,
                seriousness_level="High",
                estimated_days=3,
                cost_added=2000,
                status="In Service",
            )
            db.add(service_record)
    else:
        # If car is not damaged and status was 'In Service' from a prior state,
        # we leave actual service completion logic to the Service cron/job.
        if not item.status:
            item.status = "Available"

    db.commit()
    db.refresh(item)
    return item

